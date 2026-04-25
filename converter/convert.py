#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SplitHub Catalog Converter
Excel (мастерфайл) -> products.js + deploy.zip
"""

import sys
import os
import json
import shutil
import zipfile
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl не установлен. Запустите: pip install openpyxl")
    sys.exit(1)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

BASE_DIR    = Path(__file__).parent
PROJECT_DIR = BASE_DIR.parent          # корень проекта splithub/
TIMESTAMP   = datetime.now().strftime("%Y%m%d_%H%M%S")


# ── helpers ────────────────────────────────────────────────

def load_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)

def safe_int(v, default=0):
    try:
        return int(float(str(v).replace(",", ".").replace(" ", "").replace("\xa0", "")))
    except (ValueError, TypeError):
        return default

def clean(v):
    if v is None:
        return ""
    return " ".join(str(v).split())


# ── Converter ──────────────────────────────────────────────

class Converter:
    def __init__(self):
        self.settings = load_json(BASE_DIR / "config" / "settings.json")
        self.mapping  = load_json(BASE_DIR / "config" / "mapping.json")

        self.input_dir  = BASE_DIR / self.settings["input_dir"]
        self.images_dir = BASE_DIR / self.settings["images_dir"]
        self.out_dir    = BASE_DIR / self.settings["output_dir"]
        self.logs_dir   = BASE_DIR / self.settings["logs_dir"]
        self.backup_dir = BASE_DIR / self.settings["backup_dir"]
        self.dry_run    = self.settings.get("dry_run", False)

        self.errors   = []
        self.warnings = []
        self.products = []
        self.missing_photos = []

    # ── Excel reading ────────────────────────────────────

    def find_excel(self):
        """Найти .xlsx в input/ — берём самый свежий."""
        files = sorted(
            list(self.input_dir.glob("*.xlsx")) + list(self.input_dir.glob("*.xls")),
            key=lambda f: f.stat().st_mtime, reverse=True
        )
        if not files:
            print("ERROR: .xlsx файл не найден в папке converter/input/")
            sys.exit(1)
        if len(files) > 1:
            self.warnings.append(f"Найдено {len(files)} файлов, используется последний: {files[0].name}")
        return files[0]

    def read_excel(self, path):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        sheet_name = self.settings.get("sheet_name")
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[0]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("ERROR: Excel пустой")
            sys.exit(1)

        header = [clean(c) for c in rows[0]]
        col_idx = {name: i for i, name in enumerate(header) if name}

        # excel_col → js_field
        field_map = {}
        for excel_col, js_field in self.mapping.items():
            if excel_col in col_idx:
                field_map[col_idx[excel_col]] = js_field
            else:
                self.warnings.append(f"Колонка '{excel_col}' не найдена в Excel")

        return rows[1:], field_map

    # ── Row processing ───────────────────────────────────

    def process_row(self, row_values, field_map, row_num):
        raw = {}
        for col_i, js_field in field_map.items():
            raw[js_field] = row_values[col_i] if col_i < len(row_values) else None

        # Пропустить пустые строки
        if all(v is None or clean(v) == "" for v in raw.values()):
            return None

        # Пропустить неактивные (active != 1)
        active = clean(raw.get("active", "1"))
        if active not in ("1", "yes", "true", "да"):
            return None

        row_errors = []

        def get(field, default=""):
            return clean(raw.get(field, default) or "")

        # BTU: привести к двузначному строковому виду
        btu_raw = get("btu")
        try:
            btu_val = str(int(float(btu_raw))).zfill(2) if btu_raw and btu_raw not in ("-", "") else btu_raw
        except (ValueError, TypeError):
            btu_val = btu_raw

        price = safe_int(raw.get("price"), 0)
        if price <= 0:
            row_errors.append(f"Строка {row_num}: цена = {raw.get('price')} (должна быть > 0)")

        # Проверка обязательных полей
        required = self.settings.get("required_fields", [])
        field_vals = {
            "sku": get("sku"), "brand": get("brand"), "model": get("model"),
            "group": get("group"), "stock": get("stock"),
            "stockLabel": get("stockLabel"), "descShort": get("descShort"),
        }
        for req in required:
            if req in field_vals and not field_vals.get(req):
                row_errors.append(f"Строка {row_num}: обязательное поле '{req}' пусто")

        # Валидация группы
        valid_groups = self.settings.get("valid_groups", [])
        grp = get("group")
        if valid_groups and grp and grp not in valid_groups:
            self.warnings.append(f"Строка {row_num}: неизвестная группа '{grp}'")

        # Валидация stock_status
        valid_stocks = self.settings.get("valid_stock_statuses", [])
        stk = get("stock")
        if valid_stocks and stk and stk not in valid_stocks:
            self.warnings.append(f"Строка {row_num}: неизвестный stock '{stk}'")

        # Benefits
        sep = self.settings.get("benefits_separator", "|")
        benefits_raw = get("benefits")
        benefits = [b.strip() for b in benefits_raw.split(sep) if b.strip()] if benefits_raw else []

        # Photo
        photo = get("photo")
        if photo:
            found = self._find_photo(photo)
            if not found:
                self.missing_photos.append({"row": row_num, "sku": get("sku"), "photo": photo})
                mode = self.settings.get("missing_photo_mode", "warn")
                msg = f"Строка {row_num}: фото '{photo}' не найдено в converter/images/"
                if mode == "error":
                    row_errors.append(msg)
                else:
                    self.warnings.append(msg)

        self.errors.extend(row_errors)
        if row_errors:
            return None

        product = {
            "id":         get("id"),
            "sku":        get("sku"),
            "brandCode":  get("brandCode"),
            "brand":      get("brand"),
            "series":     get("series"),
            "model":      get("model"),
            "group":      get("group"),
            "factory":    get("factory"),
            "color":      get("color") or "white",
            "btu":        btu_val,
            "area":       safe_int(raw.get("area"), 0),
            "price":      price,
            "stock":      get("stock"),
            "stockLabel": get("stockLabel"),
            "descShort":  get("descShort"),
            "cardBenef":  get("cardBenef"),
            "benefits":   benefits,
            "compressor": get("compressor"),
            "freon":      get("freon"),
            "photo":      photo,
            "_sortOrder": safe_int(raw.get("sortOrder"), 999),
        }
        return product

    def _find_photo(self, name):
        if not name or not self.images_dir.exists():
            return None
        p = self.images_dir / name
        if p.exists():
            return name
        for f in self.images_dir.iterdir():
            if f.name.lower() == name.lower():
                return f.name
        return None

    # ── Generate JS ──────────────────────────────────────

    def generate_js(self):
        # Убираем служебное поле _sortOrder
        clean_products = []
        for p in self.products:
            cp = {k: v for k, v in p.items() if k != "_sortOrder"}
            clean_products.append(cp)

        lines = ["var PRODUCTS = ["]
        for i, p in enumerate(clean_products):
            comma = "," if i < len(clean_products) - 1 else ""
            lines.append("  " + json.dumps(p, ensure_ascii=False) + comma)
        lines.append("];")
        return "\n".join(lines) + "\n"

    # ── Copy photos ──────────────────────────────────────

    def copy_photos(self):
        photo_out = self.out_dir / "assets" / "img" / "products"
        photo_out.mkdir(parents=True, exist_ok=True)
        copied = 0
        for p in self.products:
            photo = p.get("photo", "")
            if not photo:
                continue
            found = self._find_photo(photo)
            if found:
                src = self.images_dir / found
                dst = photo_out / found
                if src.exists() and not dst.exists():
                    shutil.copy2(src, dst)
                    copied += 1
        return copied

    # ── Build deploy.zip ─────────────────────────────────

    def build_zip(self):
        """
        Собирает deploy.zip из:
        - products.js (сгенерированный)
        - все остальные файлы проекта (index.html, send.php, api/, db/, assets/, .htaccess)
        Исключает: .git, converter/, api/debug.php, *.sqlite
        """
        zip_path = self.out_dir / f"deploy_{TIMESTAMP}.zip"

        EXCLUDE_DIRS  = {".git", "converter", ".agents", ".claude", "node_modules"}
        EXCLUDE_FILES = {"api/debug.php", "db/splithub.sqlite"}
        EXCLUDE_EXT   = {".sqlite", ".db"}

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            # 1. products.js — берём сгенерированный из out/
            products_js = self.out_dir / "products.js"
            if products_js.exists():
                zf.write(products_js, "products.js")

            # 2. Фото из out/assets/img/products/
            photo_out = self.out_dir / "assets" / "img" / "products"
            if photo_out.exists():
                for f in photo_out.iterdir():
                    if f.is_file():
                        zf.write(f, f"assets/img/products/{f.name}")

            # 3. Все остальные файлы проекта (кроме products.js — уже взяли)
            for item in PROJECT_DIR.rglob("*"):
                if not item.is_file():
                    continue

                # Проверяем исключения по части пути
                parts = item.relative_to(PROJECT_DIR).parts
                if any(p in EXCLUDE_DIRS for p in parts):
                    continue
                if item.suffix in EXCLUDE_EXT:
                    continue

                rel = item.relative_to(PROJECT_DIR).as_posix()
                if rel in EXCLUDE_FILES:
                    continue
                if rel == "products.js":
                    continue  # уже добавили сгенерированный

                zf.write(item, rel)

        return zip_path

    # ── Backup ───────────────────────────────────────────

    def backup_existing(self):
        src = self.out_dir / "products.js"
        if src.exists():
            dst = self.backup_dir / TIMESTAMP
            dst.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dst / "products.js")
            print(f"  Резервная копия: converter/backup/{TIMESTAMP}/")

    # ── Write logs ───────────────────────────────────────

    def write_logs(self, excel_name):
        self.logs_dir.mkdir(exist_ok=True)
        log_path = self.logs_dir / f"run_{TIMESTAMP}.txt"
        with open(log_path, "w", encoding="utf-8") as f:
            f.write(f"=== SplitHub Converter {TIMESTAMP} ===\n")
            f.write(f"Файл: {excel_name}\n\n")
            f.write(f"Товаров: {len(self.products)}\n")
            f.write(f"Ошибок: {len(self.errors)}\n")
            f.write(f"Предупреждений: {len(self.warnings)}\n\n")
            if self.errors:
                f.write("=== ОШИБКИ ===\n")
                for e in self.errors:
                    f.write(f"  [ERR] {e}\n")
            if self.warnings:
                f.write("\n=== ПРЕДУПРЕЖДЕНИЯ ===\n")
                for w in self.warnings:
                    f.write(f"  [WARN] {w}\n")
            if self.missing_photos:
                f.write("\n=== ФОТО НЕ НАЙДЕНЫ ===\n")
                for mp in self.missing_photos:
                    f.write(f"  строка {mp['row']}  sku={mp['sku']}  фото={mp['photo']}\n")
        return log_path

    # ── Run ──────────────────────────────────────────────

    def run(self):
        print(f"\n{'='*52}")
        print("  SplitHub Catalog Converter")
        print(f"{'='*52}\n")

        # 1. Найти Excel
        excel_path = self.find_excel()
        print(f"Файл:  {excel_path.name}")

        # 2. Читать Excel
        print("Читаю Excel...")
        data_rows, field_map = self.read_excel(excel_path)
        print(f"  Строк данных: {len(data_rows)}")

        # 3. Обработать строки
        print("Обрабатываю товары...")
        skipped = 0
        for i, row in enumerate(data_rows, start=2):
            if all(v is None for v in row):
                continue
            product = self.process_row(row, field_map, i)
            if product is None:
                skipped += 1
            else:
                self.products.append(product)

        # 4. Сортировка по sort_order → id
        self.products.sort(key=lambda p: (p["_sortOrder"], safe_int(p["id"], 9999)))

        print(f"  Включено: {len(self.products)}  /  Пропущено: {skipped}")

        # 5. Проверка дублей SKU
        seen = {}
        for p in self.products:
            if p["sku"] in seen:
                self.errors.append(f"Дубль SKU: '{p['sku']}'")
            seen[p["sku"]] = True

        # 6. Вывод ошибок / предупреждений
        if self.errors:
            print(f"\n  [!] ОШИБКИ ({len(self.errors)}):")
            for e in self.errors[:10]:
                print(f"      {e}")
            if len(self.errors) > 10:
                print(f"      ... ещё {len(self.errors)-10} (см. logs/)")

        if self.warnings:
            print(f"\n  Предупреждений: {len(self.warnings)}")
            for w in self.warnings[:5]:
                print(f"      {w}")

        # 7. Лог
        log_path = self.write_logs(excel_path.name)

        # 8. Остановить при критических ошибках (если не dry-run)
        if self.errors and not self.dry_run:
            print(f"\n[СТОП] Исправьте Excel и запустите снова.")
            print(f"       Лог: converter/logs/{log_path.name}\n")
            return False

        if self.dry_run:
            print(f"\n[DRY-RUN] products.js будет содержать {len(self.products)} товаров\n")
            return True

        # 9. Backup
        self.backup_existing()

        # 10. Записать products.js в out/
        self.out_dir.mkdir(exist_ok=True)
        js_content = self.generate_js()
        (self.out_dir / "products.js").write_text(js_content, encoding="utf-8")
        print(f"\nСгенерировано: converter/out/products.js  ({len(self.products)} товаров)")

        # 11. Скопировать products.js в корень проекта (для git)
        shutil.copy2(self.out_dir / "products.js", PROJECT_DIR / "products.js")
        print(f"Скопировано:   products.js  → корень проекта")

        # 12. Фото
        copied = self.copy_photos()
        if copied:
            print(f"Фото:          {copied} скопировано в out/assets/img/products/")
        if self.missing_photos:
            print(f"Фото не найдено: {len(self.missing_photos)} (добавьте в converter/images/)")

        # 13. Deploy ZIP
        print("\nСобираю deploy.zip...")
        zip_path = self.build_zip()
        zip_size = zip_path.stat().st_size / 1024
        print(f"Готово:        {zip_path.name}  ({zip_size:.0f} КБ)")

        print(f"\n{'='*52}")
        if self.warnings:
            print("СТАТУС: УСПЕШНО (есть предупреждения)")
        else:
            print("СТАТУС: УСПЕШНО")
        print(f"{'='*52}")
        print(f"\nЧто делать дальше:")
        print(f"  1. Проверьте converter/out/{zip_path.name}")
        print(f"  2. Распакуйте на сервер в public_html/")
        print(f"  3. Или загрузите только products.js (если меняли только каталог)\n")
        return True


def main():
    dry_run = "--dry-run" in sys.argv or "-d" in sys.argv
    conv = Converter()
    if dry_run:
        conv.dry_run = True
    ok = conv.run()
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
